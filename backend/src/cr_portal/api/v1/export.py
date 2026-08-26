from datetime import date
from io import BytesIO
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload
from cr_portal.api.deps import db_session
from cr_portal.models.bonus import BonusCalculation
from cr_portal.models.kpi import CalculationIssue
from cr_portal.models.user import User
router=APIRouter()
def parse_month(v:str)->date:
    try:y,m=map(int,v.split("-"));return date(y,m,1)
    except Exception as e:raise HTTPException(422,"month must be YYYY-MM") from e

@router.get("/excel")
async def excel(month:str=Query(...),session:AsyncSession=Depends(db_session)):
    m=parse_month(month)
    rr=await session.execute(select(BonusCalculation).options(selectinload(BonusCalculation.items)).where(BonusCalculation.month==m).order_by(BonusCalculation.employee_id,BonusCalculation.version.desc()))
    latest={}
    for x in rr.scalars().all():latest.setdefault(x.employee_id,x)
    users={u.id:u.full_name for u in (await session.execute(select(User))).scalars().all()}
    wb=Workbook();ws=wb.active;ws.title="Общий отчет"
    ws.append(["Сотрудник","Внедрение база","Тех интеграция база","Часы","Продажа","Обучение","CR Start фикс","Делимая часть","Итого","Версия"])
    for c in latest.values():ws.append([users.get(c.employee_id,str(c.employee_id)),float(c.implementation_total),float(c.tech_integration_total),float(c.support_hours),float(c.sales_total),c.training_count,float(c.cr_start_fixed_total),float(c.subtotal_dividable),float(c.total_bonus),c.version])
    d=wb.create_sheet("Детализация");d.append(["Сотрудник","Тип","Источник","База","Ставка","Количество","До делителя","2.5","Итого","Описание"])
    for c in latest.values():
        for i in c.items:d.append([users.get(c.employee_id,str(c.employee_id)),i.bonus_type,i.source_external_id,float(i.base_amount),float(i.rate),float(i.quantity),float(i.amount_before_divider),"Да" if i.divider_applied else "Нет",float(i.amount_final),i.description])
    e=wb.create_sheet("Ошибки");e.append(["Уровень","Код","Сообщение","Сотрудник","Сделка"])
    for i in (await session.execute(select(CalculationIssue).where(CalculationIssue.month==m))).scalars().all():e.append([i.severity,i.code,i.message,users.get(i.employee_id,"") if i.employee_id else "",str(i.deal_id or "")])
    b=BytesIO();wb.save(b);b.seek(0)
    return StreamingResponse(b,media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",headers={"Content-Disposition":f'attachment; filename="cr_portal_{m:%Y_%m}.xlsx"'})
